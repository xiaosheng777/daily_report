from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.core.text import normalize_text

DAILY_SHEET_NAME = "日报"
CASE_ID_KEYS = ("用例编号", "case_id", "编号", "ID", "id", "Case ID")
DEFAULT_COMPARE_FIELDS = ("描述", "测试用例设计说明", "具体步骤", "期望结果")


def load_daily_sheet_rows(source: str | Path | bytes, sheet_name: str = DAILY_SHEET_NAME) -> dict[str, dict[str, Any]]:
    """Parse the named sheet into {case_id: row_dict}. Raises ValueError if sheet missing."""
    if isinstance(source, (bytes, bytearray)):
        workbook = load_workbook(BytesIO(source), data_only=True)
    else:
        workbook = load_workbook(Path(source), data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"缺少 sheet：{sheet_name}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [_clean_header(x) for x in rows[0]]
    out: dict[str, dict[str, Any]] = {}
    for idx, values in enumerate(rows[1:], start=2):
        row = {
            (headers[i] if i < len(headers) and headers[i] else f"col_{i + 1}"): _cell_text(values[i] if i < len(values) else "")
            for i in range(max(len(headers), len(values or ())))
        }
        if not any(str(v or "").strip() for v in row.values()):
            continue
        key = case_key(row) or f"row_{idx}"
        out[key] = row
    return out


def rows_to_items(rows: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for case_id, row in rows.items():
        item = dict(row)
        item["case_id"] = case_key(row) or case_id
        items.append(item)
    return items


def case_key(row: dict[str, Any]) -> str:
    for key in CASE_ID_KEYS:
        if str(row.get(key) or "").strip():
            return str(row.get(key)).strip()
    return ""


def field_value(row: dict[str, Any], field: str) -> str:
    if field in row:
        return _cell_text(row.get(field))
    aliases = {
        "描述": ("描述", "description", "名称", "title"),
        "测试用例设计说明": ("测试用例设计说明", "设计说明", "design"),
        "具体步骤": ("具体步骤", "步骤", "steps"),
        "期望结果": ("期望结果", "expected", "预期结果"),
    }
    for alias in aliases.get(field, (field,)):
        if alias in row and str(row.get(alias) or "").strip():
            return _cell_text(row.get(alias))
    return ""


def compare_fields_equal(left: dict[str, Any], right: dict[str, Any], fields: list[str] | tuple[str, ...] | None = None) -> bool:
    compare = list(fields or DEFAULT_COMPARE_FIELDS)
    for field in compare:
        if normalize_text(field_value(left, field)) != normalize_text(field_value(right, field)):
            return False
    return True


def differing_fields(left: dict[str, Any], right: dict[str, Any], fields: list[str] | tuple[str, ...] | None = None) -> list[str]:
    compare = list(fields or DEFAULT_COMPARE_FIELDS)
    return [field for field in compare if normalize_text(field_value(left, field)) != normalize_text(field_value(right, field))]


def _clean_header(value: Any) -> str:
    text = str(value or "").replace("\r", "").strip()
    return " ".join(part.strip() for part in text.splitlines() if part.strip())


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
